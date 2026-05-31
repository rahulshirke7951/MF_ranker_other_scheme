import pandas as pd
import re
from datetime import datetime
from typing import Tuple
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class Config:
    INPUT_FILE: str = "dashboard_data.xlsx"
    OUTPUT_FILE: str = "mf_ranked_screener.xlsx"
    TREND_BONUS: float = 5.0

CONFIG = Config()

QUALITY_FILTERS = {
    "cagr_2y_min": 0.10,
    "cagr_3y_min": 0.12,
    "min_1y_return": -30.0,
}

ENGINE1_WEIGHTS = {
    "return_6m": 0.30,
    "return_3m": 0.20,
    "return_1y": 0.25,
    "return_1m": 0.25,
}

ENGINE2_WEIGHTS = {
    "return_1y": 0.25,
    "return_2y": 0.30,
    "return_3y": 0.45,
}

COMPOSITE_BLEND = {
    "engine1_momentum": 0.55,
    "engine2_quality": 0.45,
}

FILTERS = {
    "cat_level_1": "Open Ended Schemes",
    "cat_level_2": "Other Scheme",
    "plan_type": "Regular",
    "option_type": "Growth",
}

COLUMN_MAP = {
    "scheme_name": "scheme_name",
    "category": "cat_level_3",
    "amc": "amc_name",
    "return_1m": "return_30d",
    "return_3m": "return_90d",
    "return_6m": "return_180d",
    "return_1y": "return_365d",
    "return_2y": "return_730d",
    "return_3y": "return_1095d",
    "nav": "latest_nav",
}

# ══════════════════════════════════════════════════════════════════════════════
# ASSET TAGGING RULES - Priority-Ordered
# ══════════════════════════════════════════════════════════════════════════════
ASSET_TAG_RULES = [
    {"tag": "Silver", "contains": ["silver"], "priority": 1},
    {"tag": "Gold", "contains": ["gold"], "priority": 2},
    {"tag": "G-SEC", "contains": ["g-sec", "gsec"], "priority": 3},
    {"tag": "GILT", "contains": ["gilt"], "priority": 4},
    {"tag": "NASDAQ", "contains": ["nasdaq", "nq100", "nq 100"], "priority": 5},
    {"tag": "S&P 500", "contains": ["s&p 500", "s&p500", "s&p 50"], "priority": 6},
    {"tag": "International", "contains": ["overseas", "global", "hang seng", "china", "greater china", "world", "emerging market"], "priority": 7},
    {"tag": "Pharma/Healthcare", "contains": ["pharma", "healthcare", "health"], "priority": 8},
    {"tag": "Technology", "contains": ["tech", "artificial intelligence", "ai ", "fang", "digital"], "priority": 9},
    {"tag": "Commodities", "contains": ["commodit", "metal", "energy", "mining", "oil", "gas"], "priority": 10},
    {"tag": "Infrastructure", "contains": ["infra", "infrastructure"], "priority": 11},
    {"tag": "Banking/Financial", "contains": ["bank", "financial", "psu bank", "private bank"], "priority": 12},
    {"tag": "Index Fund", "contains": ["index", "nifty", "sensex", "bse", "midcap", "smallcap", "next 50", "nifty 50", "nifty50"], "priority": 13},
]

# ══════════════════════════════════════════════════════════════════════════════
# COLORS
# ══════════════════════════════════════════════════════════════════════════════
class Colors:
    TITLE_BG = "0D1117"
    TITLE_FG = "FFFFFF"
    INFO_BG = "F0F4F8"
    INFO_FG = "445566"
    BORDER = "D0D8E4"
    COL_HDR_BG = "1A3A5C"
    COL_HDR_FG = "FFFFFF"
    MOMENTUM_BG = "E8560A"
    MOMENTUM_FG = "FFFFFF"
    LONGTERM_BG = "1E6B8C"
    LONGTERM_FG = "FFFFFF"
    ENGINE_BG = "2D5016"
    ENGINE_FG = "FFFFFF"
    SIGNAL_BG = "6A1B9A"  # Purple for signals
    SIGNAL_FG = "FFFFFF"
    RANK1_BG = "FFD700"
    RANK2_BG = "E8E8E8"
    RANK3_BG = "D4956A"
    ALT_ROW = "F5F8FC"
    WHITE = "FFFFFF"
    POSITIVE = "1E7A4B"
    NEGATIVE = "C0392B"
    ENGINE1_TINT = "FFF3E0"
    ENGINE2_TINT = "E3F2FD"
    COMP_TINT = "F0FFF4"
    MOMENTUM_ONLY_BG = "FFF8E1"
    MOMENTUM_ONLY_FG = "E65100"
    MISSING_DATA_BG = "ECEFF1"
    MISSING_DATA_FG = "78909C"
    ASMP_TITLE = "0D1117"
    ASMP_BLEND = "2D5016"
    ASMP_ROW_BL = "EAFAF1"
    LEGEND_FULL = "E8F5E9"
    LEGEND_MOMENTUM = "FFF8E1"
    LEGEND_MISSING = "ECEFF1"
    CONSOLIDATED_BG = "1565C0"
    STATUS_FULL = "4CAF50"
    STATUS_MOMENTUM = "FF9800"
    STATUS_MISSING = "9E9E9E"

C = Colors()

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for col_key, value in FILTERS.items():
        actual = cols_lower.get(col_key.lower().strip())
        if actual:
            df = df[df[actual].astype(str).str.strip().str.lower() == str(value).strip().lower()]
    return df

def load_data() -> pd.DataFrame:
    sheets = pd.read_excel(CONFIG.INPUT_FILE, sheet_name=None)
    frames = [df for df in sheets.values() if "scheme_name" in df.columns]
    if not frames:
        raise ValueError("No valid sheets found with 'scheme_name' column")
    return apply_filters(pd.concat(frames, ignore_index=True))

# ══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ══════════════════════════════════════════════════════════════════════════════
def to_num(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace('%', '', regex=False).str.replace(',', '', regex=False).str.strip()
    return pd.to_numeric(s, errors='coerce')

def pct_rank(series: pd.Series) -> pd.Series:
    valid = series.dropna()
    if valid.empty:
        return series.fillna(0.0)
    ranks = series.rank(method='min', na_option='bottom')
    return (ranks - 1) / max(len(series) - 1, 1) * 100

def cagr_2y(r: float) -> float:
    if pd.isna(r) or r <= -100:
        return float('nan')
    return ((1 + float(r) / 100) ** 0.5 - 1) * 100

def assign_asset_tag(scheme_name: str) -> str:
    name_lower = scheme_name.lower()
    for rule in sorted(ASSET_TAG_RULES, key=lambda x: x.get("priority", 99)):
        for keyword in rule.get("contains", []):
            if keyword in name_lower:
                return rule["tag"]
    return "Standard Equity/Debt"

def calculate_trend_strength(row: pd.Series) -> Tuple[float, str]:
    r1m, r3m, r6m = row["_r1m"], row["_r3m"], row["_r6m"]
    if pd.isna(r1m) or pd.isna(r3m) or pd.isna(r6m):
        return 0, ""
    if r6m > r3m > r1m:
        return CONFIG.TREND_BONUS, "📈 Uptrend"
    elif r6m > r3m or r3m > r1m:
        return CONFIG.TREND_BONUS / 2, "↗️ Moderate"
    elif r6m < r3m < r1m:
        return -CONFIG.TREND_BONUS / 2, "📉 Downtrend"
    return 0, ""

def classify_data_status(row: pd.Series) -> str:
    momentum_cols = ["_r1m", "_r3m", "_r6m"]
    longterm_cols = ["_r2y_cagr", "_r3y"]
    has_momentum = not row[momentum_cols].isna().any()
    has_longterm = not row[longterm_cols].isna().any()
    if has_momentum and has_longterm:
        return "FULL"
    elif has_momentum and not has_longterm:
        return "MOMENTUM_ONLY"
    else:
        return "MISSING"

def score_funds(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_r1m"] = to_num(df[COLUMN_MAP["return_1m"]])
    df["_r3m"] = to_num(df[COLUMN_MAP["return_3m"]])
    df["_r6m"] = to_num(df[COLUMN_MAP["return_6m"]])
    df["_r1y"] = to_num(df[COLUMN_MAP["return_1y"]])
    df["_r2y_raw"] = to_num(df[COLUMN_MAP["return_2y"]])
    df["_r2y_cagr"] = df["_r2y_raw"].apply(cagr_2y)
    df["_r3y"] = to_num(df[COLUMN_MAP["return_3y"]])
    df["_cat"] = df[COLUMN_MAP["category"]].astype(str).str.strip().str.title()
    df["_asset_class"] = df[COLUMN_MAP["scheme_name"]].apply(assign_asset_tag)
    df["_data_status"] = df.apply(classify_data_status, axis=1)
    df["_has_missing_data"] = df["_data_status"] == "MISSING"
    
    mask_2y = df["_r2y_cagr"] > (QUALITY_FILTERS["cagr_2y_min"] * 100)
    mask_3y = df["_r3y"] > (QUALITY_FILTERS["cagr_3y_min"] * 100)
    mask_drawdown = df["_r1y"] > QUALITY_FILTERS["min_1y_return"]
    df["_qualifies"] = mask_2y & mask_3y & mask_drawdown & (df["_data_status"] == "FULL")
    
    trend_results = df.apply(calculate_trend_strength, axis=1)
    df["_trend_bonus"] = trend_results.apply(lambda x: x[0])
    df["_trend"] = trend_results.apply(lambda x: x[1])
    
    df["_e1"] = 0.0
    df["_e2"] = 0.0
    
    for cat in df["_cat"].unique():
        cm = df["_cat"] == cat
        cq = cm & df["_qualifies"]
        valid_m_mask = cm & (df["_data_status"].isin(["FULL", "MOMENTUM_ONLY"]))
        
        if valid_m_mask.sum() > 0:
            e1 = (pct_rank(df.loc[valid_m_mask, "_r6m"]) * ENGINE1_WEIGHTS["return_6m"] +
                  pct_rank(df.loc[valid_m_mask, "_r3m"]) * ENGINE1_WEIGHTS["return_3m"] +
                  pct_rank(df.loc[valid_m_mask, "_r1y"]) * ENGINE1_WEIGHTS["return_1y"] +
                  pct_rank(df.loc[valid_m_mask, "_r1m"]) * ENGINE1_WEIGHTS["return_1m"])
            e1 = e1 + df.loc[valid_m_mask, "_trend_bonus"]
            df.loc[valid_m_mask, "_e1"] = e1.clip(0, 100)
        
        if cq.sum() > 0:
            e2 = (pct_rank(df.loc[cq, "_r1y"]) * ENGINE2_WEIGHTS["return_1y"] +
                  pct_rank(df.loc[cq, "_r2y_cagr"]) * ENGINE2_WEIGHTS["return_2y"] +
                  pct_rank(df.loc[cq, "_r3y"]) * ENGINE2_WEIGHTS["return_3y"])
            df.loc[cq, "_e2"] = e2
    
    df["_comp"] = (df["_e1"] * COMPOSITE_BLEND["engine1_momentum"] +
                   df["_e2"] * COMPOSITE_BLEND["engine2_quality"])
    df.loc[df["_data_status"] == "MISSING", "_comp"] = -1.0
    df["_rank"] = (df.groupby("_cat")["_comp"]
                   .rank(method='min', ascending=False)
                   .astype(int))
    return df

# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL FUNCTIONS - 3 Types
# ══════════════════════════════════════════════════════════════════════════════
def momentum_signal(e1: float, data_status: str = "FULL") -> str:
    """Generate momentum-based signal from Engine 1 score."""
    if data_status == "MISSING":
        return "❌ N/A"
    if e1 >= 90:
        return "🔥 Hot"
    if e1 >= 75:
        return "⭐ Strong"
    if e1 >= 60:
        return "📈 Good"
    if e1 >= 40:
        return "➡️ Neutral"
    return "📉 Weak"

def quality_signal(e2: float, data_status: str = "FULL") -> str:
    """Generate quality-based signal from Engine 2 score."""
    if data_status == "MISSING":
        return "❌ N/A"
    if data_status == "MOMENTUM_ONLY":
        return "⏳ New Fund"
    if e2 >= 90:
        return "🏆 Elite"
    if e2 >= 75:
        return "⭐ Strong"
    if e2 >= 60:
        return "🏛️ Solid"
    if e2 >= 40:
        return "➡️ Average"
    if e2 > 0:
        return "⚠️ Below Avg"
    return "🔴 Not Qualified"

def composite_signal(comp: float, trend: str = "", e1: float = 0, e2: float = 0, data_status: str = "FULL") -> str:
    """Generate composite signal combining both engines."""
    if data_status == "MISSING":
        return "❌ Missing Data"
    if data_status == "MOMENTUM_ONLY":
        if e1 >= 80:
            return "🔥 Hot Momentum"
        elif e1 >= 60:
            return "📈 Momentum Only"
        return "⏳ New Fund"
    if comp >= 85 and "Uptrend" in str(trend):
        return "🚀 Strong Conviction"
    if comp >= 75:
        return "⭐ Strong Buy"
    if comp >= 60:
        if e1 > e2:
            return "📈 Momentum Play"
        return "🏛️ Quality Hold"
    if comp >= 55:
        return "✅ Buy"
    if comp >= 40:
        return "⚠️ Watch"
    return "🔴 Avoid"

# ══════════════════════════════════════════════════════════════════════════════
# FORMATTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def border() -> Border:
    s = Side(style='thin', color=C.BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_pct(val) -> str:
    if pd.isna(val) or val is None:
        return "—"
    try:
        return f"{float(val):+.2f}%"
    except:
        return "—"

def score_col(val) -> str:
    try:
        v = float(val)
        if v < 0:
            return C.MISSING_DATA_FG
    except:
        return C.MISSING_DATA_FG
    if v >= 75:
        return C.POSITIVE
    if v >= 50:
        return "E67E00"
    return C.NEGATIVE

def clean_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', '', str(name))[:31]

def hfont(size: int = 9, color: str = "FFFFFF", bold: bool = True) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color)

def dfont(size: int = 9, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color)

def get_row_style(row, row_num: int):
    data_status = row["_data_status"]
    rank = row["_rank"]
    if data_status == "MISSING":
        return (C.MISSING_DATA_BG, C.MISSING_DATA_FG, True, False)
    elif data_status == "MOMENTUM_ONLY":
        return (C.MOMENTUM_ONLY_BG, C.MOMENTUM_ONLY_FG, True, False)
    else:
        if rank == 1:
            return (C.RANK1_BG, "000000", False, True)
        elif rank == 2:
            return (C.RANK2_BG, "000000", False, True)
        elif rank == 3:
            return (C.RANK3_BG, "000000", False, True)
        else:
            bg = C.ALT_ROW if row_num % 2 == 0 else C.WHITE
            return (bg, "000000", False, False)

# ══════════════════════════════════════════════════════════════════════════════
# COLUMN LAYOUTS - Updated with Data Status and 3 Signals
# ══════════════════════════════════════════════════════════════════════════════
# Category Sheet: Rank, Scheme, AMC, Asset Class, Returns(6), Engines(3), Signals(3), Data Status
COL_HEADERS_CATEGORY = [
    "Rank", "Scheme Name", "AMC", "Asset Class",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal", "Data\nStatus"
]

# Summary Sheet: Rank, Asset Class, Scheme, AMC, Category, Returns(6), Engines(3), Signals(3)
COL_HEADERS_SUMMARY = [
    "Rank", "Asset Class", "Top Scheme", "AMC", "Category",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal"
]

# Consolidated Sheet: Rank, Scheme, AMC, Category, Asset Class, Returns(6), Engines(3), Signals(3), Data Status
COL_HEADERS_CONSOLIDATED = [
    "Rank", "Scheme Name", "AMC", "Category", "Asset Class",
    "1M\nReturn", "3M\nReturn", "6M\nReturn",
    "1Y\nReturn", "2Y\nCAGR", "3Y\nCAGR",
    "Engine 1\n(Momentum)", "Engine 2\n(Quality)", "Composite\nScore",
    "Momentum\nSignal", "Quality\nSignal", "Composite\nSignal", "Data\nStatus"
]

COL_WIDTHS_CATEGORY = [6, 45, 20, 16, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14, 12]
COL_WIDTHS_SUMMARY = [6, 16, 42, 20, 14, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14]
COL_WIDTHS_CONSOLIDATED = [6, 42, 20, 14, 14, 9, 9, 9, 9, 9, 9, 12, 12, 12, 12, 12, 14, 12]

# Column indices for formatting
RETURN_COLS_IDX_CAT = {5, 6, 7, 8, 9, 10}
RETURN_COLS_IDX_SUMMARY = {6, 7, 8, 9, 10, 11}
RETURN_COLS_IDX_CONSOLIDATED = {6, 7, 8, 9, 10, 11}

# Group header positions
MOMENTUM_COLS_CAT = (5, 7)
LONGTERM_COLS_CAT = (8, 10)
ENGINE_COLS_CAT = (11, 13)
SIGNAL_COLS_CAT = (14, 16)

# ══════════════════════════════════════════════════════════════════════════════
# BUILD CATEGORY SHEET - With Data Status Column and 3 Signals
# ══════════════════════════════════════════════════════════════════════════════
def build_category_sheet(wb, cat, cat_df):
    ws = wb.create_sheet(clean_name(cat))
    bd = border()
    ncols = len(COL_HEADERS_CATEGORY)
    data_start_row = 5
    data_end_row = data_start_row + len(cat_df) - 1
    status_col = ncols  # Data Status is last column (17)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"✦  {cat.upper()}  ✦"
    c.font = Font(name="Arial", bold=True, size=14, color=C.TITLE_FG)
    c.fill = fill(C.TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # DYNAMIC Info row using Data Status column (Q)
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    
    # Formula references Data Status column (Q)
    dynamic_formula = (
        f'="Dual Engine: {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'✅ Full: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"FULL")&" | '
        f'⚠️ Momentum: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"MOMENTUM_ONLY")&" | '
        f'❌ Missing: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"MISSING")&" | '
        f'Total: "&COUNTA(B{data_start_row}:B{data_end_row})'
    )
    
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Group headers row 3
    for ci in range(1, 5):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (MOMENTUM_COLS_CAT[0], MOMENTUM_COLS_CAT[1], "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (LONGTERM_COLS_CAT[0], LONGTERM_COLS_CAT[1], "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (ENGINE_COLS_CAT[0], ENGINE_COLS_CAT[1], "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (SIGNAL_COLS_CAT[0], SIGNAL_COLS_CAT[1], "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd
    
    # Data Status header (column 17)
    ws.cell(row=3, column=status_col).fill = fill(C.COL_HDR_BG)
    ws.cell(row=3, column=status_col).border = bd
    ws.row_dimensions[3].height = 18

    # Column headers row 4
    for ci, hdr in enumerate(COL_HEADERS_CATEGORY, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    # Data rows
    for i, (_, row) in enumerate(cat_df.iterrows(), data_start_row):
        rank = row["_rank"]
        data_status = row["_data_status"]
        rbg, text_color, is_italic, is_bold = get_row_style(row, i)
        
        # Value assignment based on data status
        if data_status == "MISSING":
            rank_val = "—"
            e1_val = "—"
            e2_val = "—"
            comp_val = "—"
        elif data_status == "MOMENTUM_ONLY":
            rank_val = f"({rank})"
            e1_val = round(row["_e1"], 1)
            e2_val = "N/A"
            comp_val = "—"
        else:
            rank_val = rank
            e1_val = round(row["_e1"], 1)
            e2_val = round(row["_e2"], 1)
            comp_val = round(row["_comp"], 1)
        
        # Generate 3 signals
        mom_sig = momentum_signal(row["_e1"], data_status)
        qual_sig = quality_signal(row["_e2"], data_status)
        comp_sig = composite_signal(row["_comp"], row.get("_trend", ""), row["_e1"], row["_e2"], data_status)
        
        vals = [
            rank_val,
            row.get(COLUMN_MAP["scheme_name"], "—"),
            row.get(COLUMN_MAP["amc"], "—"),
            row["_asset_class"],
            fmt_pct(row["_r1m"]), fmt_pct(row["_r3m"]), fmt_pct(row["_r6m"]),
            fmt_pct(row["_r1y"]), fmt_pct(row["_r2y_cagr"]), fmt_pct(row["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig,
            data_status
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4} else "center", vertical="center")
            
            if data_status == "MISSING":
                c.font = Font(name="Arial", italic=True, size=9, color=C.MISSING_DATA_FG)
            elif data_status == "MOMENTUM_ONLY":
                if ci in RETURN_COLS_IDX_CAT and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", italic=True, size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
                elif ci == 11:  # Engine 1
                    c.font = Font(name="Arial", bold=True, italic=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci in {14, 15, 16}:  # Signal columns
                    c.font = Font(name="Arial", bold=True, italic=True, size=8, color=C.MOMENTUM_ONLY_FG)
                elif ci == 17:  # Data Status
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MOMENTUM)
                    c.fill = fill(C.LEGEND_MOMENTUM)
                else:
                    c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
            else:  # FULL data
                if ci in RETURN_COLS_IDX_CAT and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", bold=is_bold, size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = dfont(bold=is_bold)
                elif ci == 11:  # Engine 1
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci == 12:  # Engine 2
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE2_TINT)
                elif ci == 13:  # Composite
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.COMP_TINT)
                elif ci in {14, 15, 16}:  # Signal columns
                    c.font = Font(name="Arial", bold=True, size=8)
                elif ci == 17:  # Data Status
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_FULL)
                    c.fill = fill(C.LEGEND_FULL)
                else:
                    c.font = dfont(bold=is_bold)
        
        ws.row_dimensions[i].height = 16

    # Column widths
    for ci, w in enumerate(COL_WIDTHS_CATEGORY, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD SUMMARY SHEET - BY ASSET CLASS WITH 3 SIGNALS
# ══════════════════════════════════════════════════════════════════════════════
def build_summary(wb, df_scored):
    ws = wb.create_sheet("🏆 SUMMARY", 0)
    bd = border()
    ncols = len(COL_HEADERS_SUMMARY)
    asset_classes = sorted(df_scored["_asset_class"].unique())
    data_start_row = 5
    data_end_row = data_start_row + len(asset_classes) - 1
    
    # Title
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "MF INTELLIGENCE — ASSET CLASS RANKED SUMMARY"
    c.font = Font(name="Arial", bold=True, size=15, color=C.TITLE_FG)
    c.fill = fill(C.TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # DYNAMIC Info row
    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    
    dynamic_formula = (
        f'="Top Performer per Asset Class | {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'Total Asset Classes: "&COUNTA(B{data_start_row}:B{data_end_row})&" | '
        f'⭐ Strong Buy+: "&COUNTIF(Q{data_start_row}:Q{data_end_row},"*Strong*")'
    )
    
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Group headers row 3
    for ci in range(1, 6):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (6, 8, "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (9, 11, "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (12, 14, "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (15, 17, "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd
    ws.row_dimensions[3].height = 18

    # Column headers row 4
    for ci, hdr in enumerate(COL_HEADERS_SUMMARY, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    # Data rows
    rank_counter = 1
    for i, asset_class in enumerate(asset_classes, data_start_row):
        asset_df = df_scored[df_scored["_asset_class"] == asset_class].copy()
        asset_df = asset_df.sort_values(
            ["_data_status", "_comp"], 
            ascending=[True, False],
            key=lambda x: x.map({"FULL": 0, "MOMENTUM_ONLY": 1, "MISSING": 2}) if x.name == "_data_status" else x
        )
        
        if asset_df.empty:
            continue
        
        top = asset_df.iloc[0]
        data_status = top["_data_status"]
        rbg, text_color, is_italic, is_bold = get_row_style(top, i)
        
        if data_status == "MISSING":
            rank_val = "—"
            e1_val = "—"
            e2_val = "—"
            comp_val = "—"
        elif data_status == "MOMENTUM_ONLY":
            rank_val = f"({rank_counter})"
            e1_val = round(top["_e1"], 1)
            e2_val = "N/A"
            comp_val = "—"
            rank_counter += 1
        else:
            rank_val = rank_counter
            e1_val = round(top["_e1"], 1)
            e2_val = round(top["_e2"], 1)
            comp_val = round(top["_comp"], 1)
            rank_counter += 1

        # Generate 3 signals
        mom_sig = momentum_signal(top["_e1"], data_status)
        qual_sig = quality_signal(top["_e2"], data_status)
        comp_sig = composite_signal(top["_comp"], top.get("_trend", ""), top["_e1"], top["_e2"], data_status)

        vals = [
            rank_val,
            asset_class,
            top.get(COLUMN_MAP["scheme_name"], "—"),
            top.get(COLUMN_MAP["amc"], "—"),
            top["_cat"],
            fmt_pct(top["_r1m"]), fmt_pct(top["_r3m"]), fmt_pct(top["_r6m"]),
            fmt_pct(top["_r1y"]), fmt_pct(top["_r2y_cagr"]), fmt_pct(top["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4, 5} else "center", vertical="center")
            
            if data_status == "MISSING":
                c.font = Font(name="Arial", italic=True, size=9, color=C.MISSING_DATA_FG)
            elif data_status == "MOMENTUM_ONLY":
                if ci in RETURN_COLS_IDX_SUMMARY and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", italic=True, size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
                elif ci == 12:  # Engine 1
                    c.font = Font(name="Arial", bold=True, italic=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci in {15, 16, 17}:  # Signals
                    c.font = Font(name="Arial", bold=True, italic=True, size=8, color=C.MOMENTUM_ONLY_FG)
                else:
                    c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
            else:
                if ci in RETURN_COLS_IDX_SUMMARY and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = dfont()
                elif ci == 12:  # Engine 1
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci == 13:  # Engine 2
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE2_TINT)
                elif ci == 14:  # Composite
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.COMP_TINT)
                elif ci in {15, 16, 17}:  # Signals
                    c.font = Font(name="Arial", bold=True, size=8)
                else:
                    c.font = dfont()
        
        ws.row_dimensions[i].height = 18

    for ci, w in enumerate(COL_WIDTHS_SUMMARY, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD CONSOLIDATED SHEET - With Data Status and 3 Signals
# ══════════════════════════════════════════════════════════════════════════════
def build_consolidated_sheet(wb, df_scored):
    ws = wb.create_sheet("📊 CONSOLIDATED")
    bd = border()
    ncols = len(COL_HEADERS_CONSOLIDATED)
    
    df_all = df_scored.copy()
    df_all = df_all.sort_values(
        ["_data_status", "_comp"], 
        ascending=[True, False],
        key=lambda x: x.map({"FULL": 0, "MOMENTUM_ONLY": 1, "MISSING": 2}) if x.name == "_data_status" else x
    ).reset_index(drop=True)
    
    data_start_row = 5
    data_end_row = data_start_row + len(df_all) - 1
    status_col = ncols  # Data Status column (18)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = "📊 CONSOLIDATED VIEW — ALL FUNDS RANKED"
    c.font = Font(name="Arial", bold=True, size=15, color=C.TITLE_FG)
    c.fill = fill(C.CONSOLIDATED_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # DYNAMIC Info row using Data Status column (R)
    e1_pct = int(COMPOSITE_BLEND["engine1_momentum"] * 100)
    e2_pct = int(COMPOSITE_BLEND["engine2_quality"] * 100)
    
    dynamic_formula = (
        f'="All Funds Consolidated | {e1_pct}% Momentum + {e2_pct}% Quality | '
        f'✅ Full: "&COUNTIF(R{data_start_row}:R{data_end_row},"FULL")&" | '
        f'⚠️ Momentum: "&COUNTIF(R{data_start_row}:R{data_end_row},"MOMENTUM_ONLY")&" | '
        f'❌ Missing: "&COUNTIF(R{data_start_row}:R{data_end_row},"MISSING")&" | '
        f'Total: "&COUNTA(B{data_start_row}:B{data_end_row})'
    )
    
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = dynamic_formula
    c.font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    c.fill = fill(C.INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Group headers row 3
    for ci in range(1, 6):
        ws.cell(row=3, column=ci).fill = fill(C.COL_HDR_BG)
        ws.cell(row=3, column=ci).border = bd

    for g_start, g_end, label, bg, fg in [
        (6, 8, "◄  Momentum  ►", C.MOMENTUM_BG, C.MOMENTUM_FG),
        (9, 11, "◄  Long-Term  ►", C.LONGTERM_BG, C.LONGTERM_FG),
        (12, 14, "◄  Engine Scores  ►", C.ENGINE_BG, C.ENGINE_FG),
        (15, 17, "◄  Signals  ►", C.SIGNAL_BG, C.SIGNAL_FG),
    ]:
        ws.merge_cells(f"{get_column_letter(g_start)}3:{get_column_letter(g_end)}3")
        cell = ws.cell(row=3, column=g_start, value=label)
        cell.font = hfont(color=fg)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(g_start, g_end + 1):
            ws.cell(row=3, column=ci).fill = fill(bg)
            ws.cell(row=3, column=ci).border = bd
    
    # Data Status header
    ws.cell(row=3, column=status_col).fill = fill(C.COL_HDR_BG)
    ws.cell(row=3, column=status_col).border = bd
    ws.row_dimensions[3].height = 18

    # Column headers row 4
    for ci, hdr in enumerate(COL_HEADERS_CONSOLIDATED, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font = hfont()
        c.fill = fill(C.COL_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bd
    ws.row_dimensions[4].height = 28

    # Data rows - Global ranking
    global_rank = 1
    for i, (_, row) in enumerate(df_all.iterrows(), data_start_row):
        data_status = row["_data_status"]
        
        # Row styling
        if data_status == "MISSING":
            rbg = C.MISSING_DATA_BG
            is_italic = True
            is_bold = False
        elif data_status == "MOMENTUM_ONLY":
            rbg = C.MOMENTUM_ONLY_BG
            is_italic = True
            is_bold = False
        else:
            if global_rank == 1:
                rbg = C.RANK1_BG
                is_bold = True
            elif global_rank == 2:
                rbg = C.RANK2_BG
                is_bold = True
            elif global_rank == 3:
                rbg = C.RANK3_BG
                is_bold = True
            elif global_rank <= 10:
                rbg = C.COMP_TINT
                is_bold = False
            else:
                rbg = C.ALT_ROW if i % 2 == 0 else C.WHITE
                is_bold = False
            is_italic = False
        
        # Value assignment
        if data_status == "MISSING":
            rank_val = "—"
            e1_val = "—"
            e2_val = "—"
            comp_val = "—"
        elif data_status == "MOMENTUM_ONLY":
            rank_val = f"({global_rank})"
            e1_val = round(row["_e1"], 1)
            e2_val = "N/A"
            comp_val = "—"
            global_rank += 1
        else:
            rank_val = global_rank
            e1_val = round(row["_e1"], 1)
            e2_val = round(row["_e2"], 1)
            comp_val = round(row["_comp"], 1)
            global_rank += 1
        
        # Generate 3 signals
        mom_sig = momentum_signal(row["_e1"], data_status)
        qual_sig = quality_signal(row["_e2"], data_status)
        comp_sig = composite_signal(row["_comp"], row.get("_trend", ""), row["_e1"], row["_e2"], data_status)
        
        vals = [
            rank_val,
            row.get(COLUMN_MAP["scheme_name"], "—"),
            row.get(COLUMN_MAP["amc"], "—"),
            row["_cat"],
            row["_asset_class"],
            fmt_pct(row["_r1m"]), fmt_pct(row["_r3m"]), fmt_pct(row["_r6m"]),
            fmt_pct(row["_r1y"]), fmt_pct(row["_r2y_cagr"]), fmt_pct(row["_r3y"]),
            e1_val, e2_val, comp_val,
            mom_sig, qual_sig, comp_sig,
            data_status
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.border = bd
            c.fill = fill(rbg)
            c.alignment = Alignment(horizontal="left" if ci in {2, 3, 4, 5} else "center", vertical="center")
            
            if data_status == "MISSING":
                c.font = Font(name="Arial", italic=True, size=9, color=C.MISSING_DATA_FG)
            elif data_status == "MOMENTUM_ONLY":
                if ci in RETURN_COLS_IDX_CONSOLIDATED and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", italic=True, size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
                elif ci == 12:  # Engine 1
                    c.font = Font(name="Arial", bold=True, italic=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci in {15, 16, 17}:  # Signals
                    c.font = Font(name="Arial", bold=True, italic=True, size=8, color=C.MOMENTUM_ONLY_FG)
                elif ci == 18:  # Data Status
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_MOMENTUM)
                    c.fill = fill(C.LEGEND_MOMENTUM)
                else:
                    c.font = Font(name="Arial", italic=True, size=9, color=C.MOMENTUM_ONLY_FG)
            else:
                if ci in RETURN_COLS_IDX_CONSOLIDATED and isinstance(val, str) and val != "—":
                    try:
                        num = float(val.replace('%', ''))
                        c.font = Font(name="Arial", bold=is_bold, size=9, color=C.POSITIVE if num >= 0 else C.NEGATIVE)
                    except:
                        c.font = dfont(bold=is_bold)
                elif ci == 12:  # Engine 1
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE1_TINT)
                elif ci == 13:  # Engine 2
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.ENGINE2_TINT)
                elif ci == 14:  # Composite
                    c.font = Font(name="Arial", bold=True, size=9, color=score_col(val))
                    c.fill = fill(C.COMP_TINT)
                elif ci in {15, 16, 17}:  # Signals
                    c.font = Font(name="Arial", bold=True, size=8)
                elif ci == 18:  # Data Status
                    c.font = Font(name="Arial", bold=True, size=8, color=C.STATUS_FULL)
                    c.fill = fill(C.LEGEND_FULL)
                else:
                    c.font = dfont(bold=is_bold)
        
        ws.row_dimensions[i].height = 16

    for ci, w in enumerate(COL_WIDTHS_CONSOLIDATED, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{data_end_row}"
    ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# BUILD ASSUMPTIONS SHEET - Updated Signal Guide
# ══════════════════════════════════════════════════════════════════════════════
def build_assumptions(wb, df_scored):
    ws = wb.create_sheet("📋 ASSUMPTIONS", 1)
    bd = border()

    def section(row, title, bg, fg="FFFFFF", ncols=3):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.font = Font(name="Arial", bold=True, size=11, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        return row + 1

    def kv_row(row, key, val, desc="", row_bg="FFFFFF"):
        for ci, text in [(1, key), (2, val), (3, desc)]:
            c = ws.cell(row=row, column=ci, value=text)
            c.font = Font(name="Arial", size=9)
            c.fill = fill(row_bg)
            c.border = bd
            c.alignment = Alignment(horizontal="left" if ci in {1, 3} else "center", vertical="center")
        ws.row_dimensions[row].height = 15
        return row + 1

    def col_headers(row, headers, bg):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill = fill(bg)
            c.border = bd
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16
        return row + 1

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "DUAL ENGINE MODEL — ASSUMPTIONS & METHODOLOGY"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill(C.ASMP_TITLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 3-Signal System Enabled"
    ws["A2"].font = Font(name="Arial", italic=True, size=8.5, color=C.INFO_FG)
    ws["A2"].fill = fill(C.INFO_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 14

    r = 4
    
    # Data Status Legend
    r = section(r, "📊  DATA AVAILABILITY LEGEND", C.ASMP_BLEND)
    r = col_headers(r, ["Status", "Visual Style", "Description"], C.ASMP_BLEND)
    r = kv_row(r, "✅ FULL", "White/Ranked Colors", "All returns (1M-3Y) available - fully scored", C.LEGEND_FULL)
    r = kv_row(r, "⚠️ MOMENTUM_ONLY", "Light Amber + Italic", "Has 1M/3M/6M but missing 2Y/3Y CAGR", C.LEGEND_MOMENTUM)
    r = kv_row(r, "❌ MISSING", "Light Gray + Italic", "Missing critical short-term data", C.LEGEND_MISSING)
    
    r += 1
    
    # Momentum Signal Guide
    r = section(r, "📈  MOMENTUM SIGNAL (Engine 1 Based)", C.MOMENTUM_BG)
    r = col_headers(r, ["Signal", "E1 Score Range", "Meaning"], C.MOMENTUM_BG)
    r = kv_row(r, "🔥 Hot", "90+", "Exceptional short-term momentum", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong", "75-89", "Strong momentum performance", C.ASMP_ROW_BL)
    r = kv_row(r, "📈 Good", "60-74", "Good momentum, above average", C.ASMP_ROW_BL)
    r = kv_row(r, "➡️ Neutral", "40-59", "Average momentum", C.ASMP_ROW_BL)
    r = kv_row(r, "📉 Weak", "<40", "Below average momentum", C.ASMP_ROW_BL)
    
    r += 1
    
    # Quality Signal Guide
    r = section(r, "🏛️  QUALITY SIGNAL (Engine 2 Based)", C.LONGTERM_BG)
    r = col_headers(r, ["Signal", "E2 Score Range", "Meaning"], C.LONGTERM_BG)
    r = kv_row(r, "🏆 Elite", "90+", "Top-tier long-term quality", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong", "75-89", "Strong quality metrics", C.ASMP_ROW_BL)
    r = kv_row(r, "🏛️ Solid", "60-74", "Solid fundamentals", C.ASMP_ROW_BL)
    r = kv_row(r, "➡️ Average", "40-59", "Average quality", C.ASMP_ROW_BL)
    r = kv_row(r, "⚠️ Below Avg", "1-39", "Below average quality", C.ASMP_ROW_BL)
    r = kv_row(r, "🔴 Not Qualified", "0", "Did not pass quality gates", C.ASMP_ROW_BL)
    r = kv_row(r, "⏳ New Fund", "N/A", "Insufficient history for quality scoring", C.LEGEND_MOMENTUM)
    
    r += 1
    
    # Composite Signal Guide
    r = section(r, "🎯  COMPOSITE SIGNAL (Combined)", C.ENGINE_BG)
    r = col_headers(r, ["Signal", "Score/Condition", "Meaning"], C.ENGINE_BG)
    r = kv_row(r, "🚀 Strong Conviction", "85+ with Uptrend", "Top performer with positive trend", C.ASMP_ROW_BL)
    r = kv_row(r, "⭐ Strong Buy", "75-84", "High composite - recommended", C.ASMP_ROW_BL)
    r = kv_row(r, "📈 Momentum Play", "60-74 (E1>E2)", "Strong short-term focus", C.ASMP_ROW_BL)
    r = kv_row(r, "🏛️ Quality Hold", "60-74 (E2>E1)", "Strong fundamentals focus", C.ASMP_ROW_BL)
    r = kv_row(r, "✅ Buy", "55-59", "Moderate buy signal", C.ASMP_ROW_BL)
    r = kv_row(r, "⚠️ Watch", "40-54", "Monitor for improvement", C.ASMP_ROW_BL)
    r = kv_row(r, "🔴 Avoid", "<40", "Below threshold", C.ASMP_ROW_BL)
    r = kv_row(r, "🔥 Hot Momentum", "E1≥80 (Momentum Only)", "Strong momentum, no history", C.LEGEND_MOMENTUM)
    r = kv_row(r, "⏳ New Fund", "Momentum Only", "Insufficient history", C.LEGEND_MOMENTUM)
    
    r += 1
    
    # Data Stats
    r = section(r, "📈  DATA STATISTICS (Live from Consolidated)", C.ASMP_BLEND)
    r = col_headers(r, ["Metric", "Count", "Percentage"], C.ASMP_BLEND)
    
    total_formula = '=COUNTA(\'📊 CONSOLIDATED\'!B5:B10000)'
    full_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"FULL")'
    momentum_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"MOMENTUM_ONLY")'
    missing_formula = '=COUNTIF(\'📊 CONSOLIDATED\'!R5:R10000,"MISSING")'
    
    for ci, (key, formula, pct_formula) in enumerate([
        ("Total Funds", total_formula, "100%"),
        ("✅ Full Data", full_formula, f'=TEXT({full_formula}/{total_formula},"0.0%")'),
        ("⚠️ Momentum Only", momentum_formula, f'=TEXT({momentum_formula}/{total_formula},"0.0%")'),
        ("❌ Missing Data", missing_formula, f'=TEXT({missing_formula}/{total_formula},"0.0%")'),
    ]):
        bg_color = C.ASMP_ROW_BL if ci == 0 else (C.LEGEND_FULL if ci == 1 else (C.LEGEND_MOMENTUM if ci == 2 else C.LEGEND_MISSING))
        
        c1 = ws.cell(row=r, column=1, value=key)
        c1.font = Font(name="Arial", size=9)
        c1.fill = fill(bg_color)
        c1.border = bd
        c1.alignment = Alignment(horizontal="left", vertical="center")
        
        c2 = ws.cell(row=r, column=2, value=formula)
        c2.font = Font(name="Arial", size=9, bold=True)
        c2.fill = fill(bg_color)
        c2.border = bd
        c2.alignment = Alignment(horizontal="center", vertical="center")
        
        c3 = ws.cell(row=r, column=3, value=pct_formula if ci > 0 else "100%")
        c3.font = Font(name="Arial", size=9)
        c3.fill = fill(bg_color)
        c3.border = bd
        c3.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[r].height = 15
        r += 1
    
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 48

# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTOR
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("🚀 Loading data matrix...")
    df = load_data()
    
    print("⚙️ Executing Scoring Engine with 3-Signal System...")
    df_scored = score_funds(df)
    
    print(f"\n📊 Data Status Summary:")
    print(f"   ✅ Full Data: {len(df_scored[df_scored['_data_status'] == 'FULL'])}")
    print(f"   ⚠️ Momentum Only: {len(df_scored[df_scored['_data_status'] == 'MOMENTUM_ONLY'])}")
    print(f"   ❌ Missing Data: {len(df_scored[df_scored['_data_status'] == 'MISSING'])}")
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    categories = sorted(df_scored["_cat"].unique())
    
    print("\n📊 Building Asset Class Summary Sheet...")
    build_summary(wb, df_scored)
    
    print("📋 Building Assumptions Sheet...")
    build_assumptions(wb, df_scored)
    
    print("📁 Generating Category Sheets...")
    for cat in categories:
        cat_df = df_scored[df_scored["_cat"] == cat].sort_values("_rank")
        if not cat_df.empty:
            build_category_sheet(wb, cat, cat_df)
    
    print("📊 Building Consolidated Sheet...")
    build_consolidated_sheet(wb, df_scored)
    
    wb.save(CONFIG.OUTPUT_FILE)
    print(f"\n✅ Done! Output: '{CONFIG.OUTPUT_FILE}'")
    print(f"   📊 Summary: Asset Class rankings with 3 signals")
    print(f"   📋 Assumptions: With 3-signal guide")
    print(f"   📁 Category Sheets: Data Status + 3 Signals")
    print(f"   📊 CONSOLIDATED: All {len(df_scored)} funds with full analysis")

if __name__ == "__main__":
    main()
